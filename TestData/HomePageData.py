import openpyxl


class HomepageData:
    test_HomePage_data = [
        {"firstname": "Rahul", "email": "Shetty@yopmail.com", "password": "Test@123", "gender": "Male"}
        , {"firstname": "Akash", "email": "Akash@yopmail.com", "password": "Test@123", "gender": "Female"}]

    @staticmethod
    def getTestData(test_Case_name):
        Dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\USER\\PycharmProjects\\PythonAutomationFramework\\TestData\\TestData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_Case_name:

                for j in range(1, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=2, column=j).value
        return[Dict]
